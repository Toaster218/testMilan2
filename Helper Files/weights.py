from parser_constants import QUANTITY_PATTERN
from excel_utils import get_last_used_row_col, cell_to_float
from file_utils import get_output_dir
from parser_utils import get_inner_qty_sku, get_product_category_or_brand, get_order_ship_price, get_total_price
from sku_mapping import SKUMapping
from datetime import datetime
from forex import Forex
import openpyxl
import logging
import os

# GLOBAL VARIABLES
WB_NAME = 'WEIGHTS.xlsx'
WEIGHT_WB_PATH = os.path.join(get_output_dir(client_file=False), WB_NAME)
SKU_MAPPING_WB_NAME = 'Amazon SKU Mapping.xlsx'
SKU_MAPPING_WB_PATH = os.path.join(get_output_dir(client_file=False), SKU_MAPPING_WB_NAME)


class OrderData():
    '''adds data to each order in passed orders list. Class assumes workbooks WEIGHTS.xlsx and
    'Amazon SKU Mapping.xlsx' are in Helper Files folder and its data integrity, fixed headers are in place.
    
    Main methods:
    add_orders_data() - adds category, brand, vmdoption, weight to order dict keys
    export_unmapped_skus() - writes unmatched/unmapped skus to txt file    

    Arguments:
    orders: list of order dicts
    sales_channel: str
    proxy_keys: dict'''

    def __init__(self, orders:list, sales_channel:str, proxy_keys:dict):
        self.sales_channel = sales_channel
        self.proxy_keys = proxy_keys
        self.pattern = QUANTITY_PATTERN[sales_channel]
        self.fx = Forex()
        self.orders = self.__init_default(orders)
        
        self.weight_data = self._parse_weights_wb()
        if self.sales_channel != 'Etsy':    
            self.sku_mapping = SKUMapping(SKU_MAPPING_WB_PATH).read_sku_mapping_to_dict()

        self.no_matching_skus = []
        self.invalid_weight_orders = 0

    def __init_default(self, orders:list) -> list:
        '''adds some default keys to each order'''
        for order in orders:
            order['tracked'], order['skip_service_selection'] = False, False
            order['shipping_service'] = ''
            order_value = get_total_price(order, self.sales_channel, return_as_float=True)
            order['total-eur'] = self.fx.convert_to_eur(order_value, order[self.proxy_keys['currency']])
        return orders

    def _parse_weights_wb(self) -> dict:
        '''returns weights data as dict from reading excel workbook'''
        ws = self._get_weight_ws()
        ws_limits = get_last_used_row_col(ws)
        weight_data = self._get_ws_data(ws, ws_limits)
        self.wb.close()
        return weight_data

    def _get_weight_ws(self):
        '''returns ws object'''
        self.wb = openpyxl.load_workbook(WEIGHT_WB_PATH)
        ws = self.wb['Weight']
        return ws
    
    def _get_ws_data(self, ws:object, ws_limits:dict) -> dict:
        '''returns worksheet data as list of dicts, with keys corresponding to header data'''
        ws_data = {}
        max_row = ws_limits['max_row']
        max_col = ws_limits['max_col']
        for r in range(2, max_row + 1):
            row_data = {}
            for c in range(2, max_col + 1):
                header = ws.cell(row=1, column=c).value
                cell_value = cell_to_float(ws.cell(row=r, column=c).value)
                row_data[header] = cell_value
            ws_data[ws.cell(row=r, column=1).value] = row_data
        return ws_data

    def __get_order_quantity(self, order:dict) -> int:
        '''returns 'quantity-purchased' order key value in integer form'''
        return int(order[self.proxy_keys['quantity-purchased']])


    def add_orders_data(self) -> list:
        '''adds properties to each order (keys):
        -weight (order weight as float)
        -vmdoption (string)
        -brand (string)
        -category (string)
        
        new keys w/ shipping pricing update 2021.11
        -tracked (bool)
        -shipping_service
        -skip_service_selection'''
        
        for order in self.orders:
            qty_purchased = self.__get_order_quantity(order)
            skus = order[self.proxy_keys['sku']]

            # Add brand / category data to order, using first item in sku list
            order = self._add_order_brand_category_data(order, skus)

            order = self._check_tracked_status(order)

            if self._validate_calculation(qty_purchased, skus):
                order = self._calc_weight_add_data(order, qty_purchased, skus)
            else:
                order = self._add_invalid_weight_data(order)
        
        percentage_invalid = self.invalid_weight_orders / len(self.orders) * 100
        logging.info(f'{percentage_invalid:.2f}% orders contain SKU\'s that are invalid for weight calculation')
        return self.orders
    
    def _check_tracked_status(self, order:dict) -> dict:
        '''adds key 'tracked' to order dict based on country, price, shipping, items purchased'''        
        if self.sales_channel == 'Etsy':
            return self.__is_etsy_tracked(order)
        else:
            return self.__is_amazon_tracked(order)

    def __is_etsy_tracked(self, order:dict) -> dict:
        '''flips order 'tracked' bool to True if meets rules for etsy marketplace'''
        shipping_price = get_order_ship_price(order, self.proxy_keys)
        if shipping_price >= 21:            
            order['shipping_service'] = 'ups'
            order['tracked'], order['skip_service_selection'] = True, True
        elif shipping_price > 0 or order['total-eur'] > 70:
            order['tracked'] = True
        else:
            pass
        return order

    def __is_amazon_tracked(self, order:dict) -> dict:
        '''flips order 'tracked' bool to True if meets rules for amazon marketplace'''
        shipping_price = get_order_ship_price(order, self.proxy_keys)
        country = order[self.proxy_keys['ship-country']]

        print(f'calc order value in eur before proceeding. sys exit hit.')
        import sys
        sys.exit()

        # conditions for specific services:
        if shipping_price >= 20:
            order['shipping_service'] = 'ups'
            order['tracked'], order['skip_service_selection'] = True, True
        elif order['category'] == 'TAROT CARDS' and country == 'UK' and self.sales_channel == 'AmazonEU':
            order['shipping_service'] = 'etonas'
            order['tracked'], order['skip_service_selection'] = True, True
        # conditions to mark as tracked:
        elif self.sales_channel == 'AmazonCOM' or order['total-eur'] > 70 or country in ['FR', 'ES', 'IT']:
            order['tracked'] = True
        return order

    def _add_order_brand_category_data(self, order:dict, skus:list) -> dict:
        '''returns order w/ added brand, category keys (title possibly for etsy based on first sku in order)'''
        if self.sales_channel == 'Etsy':
            order = self._add_etsy_order_title(order, skus)
        title = order[self.proxy_keys['title']]
        order['brand'] = get_product_category_or_brand(title, return_brand=True)
        order['category'] = get_product_category_or_brand(title, return_brand=False)
        return order

    def _add_etsy_order_title(self, order:dict, skus:list) -> dict:
        '''adds Etsy order title to order dict'''
        for sku in skus:
            _, inner_sku = get_inner_qty_sku(sku, self.pattern)
            try:
                sku_weight_data = self.weight_data[inner_sku]
                title = sku_weight_data['Title']
                if title:
                    order['title'] = title
                    logging.debug(f'Adding title to etsy order: {title} based on inner sku: {inner_sku}')
                    return order
            except:
                continue
        # no valid title found
        order['title'] = 'Title not available'
        return order

    def _validate_calculation(self, qty_purchased:int, skus:list) -> bool:
        '''returns False if: for Etsy orders, when weight can not be calculated due to various possible combinations'''
        if self.sales_channel == 'Etsy':
            if len(skus) > 1 and qty_purchased != len(skus):
                logging.debug(f'Etsy order weights can\'t be calculated due to various possible combinations. Qty: {qty_purchased}, skus: {skus}')
                return False
        return True

    def _calc_weight_add_data(self, order:dict, qty_purchased:int, skus:list) -> dict:
        '''adds weight related data to order dict'''
        order_weight = 0.0
        package_weight = 0.0
        self.vmdoption = ''
        try:
            for sku in skus:
                inner_qty, inner_sku = get_inner_qty_sku(sku, self.pattern)

                if self.sales_channel != 'Etsy' and inner_sku not in self.weight_data:
                    # try to find sku in mapping
                    mapped_sku = self.sku_mapping[sku]
                    logging.debug(f'Found mapping match for {inner_sku}. Trying to use new (unparsed for inner) sku: {mapped_sku}')
                    inner_qty, inner_sku = get_inner_qty_sku(mapped_sku, self.pattern)
                    
                sku_weight_data = self.weight_data[inner_sku]

                sku_weight = float(sku_weight_data['Weight'])
                order_sku_weight = sku_weight * inner_qty
                
                # multuply with external order quantity only when order contains single sku
                if len(skus) == 1:
                    order_weight += order_sku_weight * qty_purchased
                else:
                    order_weight += order_sku_weight

                potential_package_weight = float(self._get_potential_package_weight(order, sku_weight_data))
                # update package weight if sku package weight is > current package weight
                if potential_package_weight > package_weight:
                    package_weight = potential_package_weight
                
                self._update_vmdoption(sku_weight_data)

            order_weight += package_weight
            order['weight'] = int(round(order_weight, 2))
            order['vmdoption'] = self.vmdoption
            return order
        except:
            return self._add_invalid_weight_data(order)

    def _get_potential_package_weight(self, order:dict, sku_weight_data:dict) -> float:
        '''returns package weight as float based on product category'''
        try:
            if order['category'] =='PLAYING CARDS' or order['category'] =='TAROT CARDS':
                return float(sku_weight_data['Package DP'])
            else:
                return float(sku_weight_data['Package LP'])
        except:
            # to fail float casting in _calc_weight_add_data
            return 'No package weight available'
    
    def _update_vmdoption(self, sku_weight_data:dict):
        '''updates self.vmdoption for order'''
        # HERE REVIEW
        # UPDATE HERE FOR VKS option, check col value
        potential_option = sku_weight_data['VMD']
        if potential_option in ['VKS', 'MKS', 'DKS']:
            if self.vmdoption == '':
                # self.vmdoption is not set (first sku in order)
                self.vmdoption = potential_option
            elif self.vmdoption == 'MKS' and potential_option == 'DKS':
                # upgrade to DKS if current option is MKS
                self.vmdoption = potential_option


    def _add_invalid_weight_data(self, order:dict) -> dict:
        '''adds invalid weight data to order dict'''
        self.invalid_weight_orders += 1
        self.no_matching_skus.append(order[self.proxy_keys['sku']])
        logging.warning(f'order: {order[self.proxy_keys["order-id"]]} cant calc weights. skus: {order[self.proxy_keys["sku"]]}')
        order['weight'] = ''
        order['vmdoption'] = ''
        return order
    
    def export_unmapped_skus(self):
        '''exports unmatched (weight or mapping) skus list to txt file'''
        date_stamp = datetime.today().strftime("%Y.%m.%d %H.%M")
        txt_path = os.path.join(get_output_dir(), f'Not matching SKUs {date_stamp}.txt')
        if self.no_matching_skus:
            with open(txt_path, 'w') as f:
                for i, sku_sublist in enumerate(self.no_matching_skus, start=1):
                    text_line = ' ,'.join(sku_sublist)
                    f.write(f'{i}. {text_line}\n')
            logging.info(f'{len(self.no_matching_skus)} skus without complete weight data or amazon mapping were written to txt file: {txt_path}')
        else:
            logging.info('All skus were matched, skipping export of self.no_matching_skus')


if __name__ == '__main__':
    pass